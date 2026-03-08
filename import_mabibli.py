#!/usr/bin/env python3
"""
Import depuis "Ma Bibliothèque" (xlsx) vers bibliotheque.json
Gère : Livres, Listes de souhaits, BD, Films, Jeux Vidéo
"""

import json
import uuid
import openpyxl
from pathlib import Path
from datetime import datetime

# ─── CONFIG ──────────────────────────────────────────────────────────────────
XLSX_FILE   = Path(__file__).parent / "MaBibliotheque.xlsx"
OUTPUT_FILE = Path(__file__).parent / "bibliotheque.json"
RESUME_MAX  = 0  # Résumé désactivé (OCR trop bruité)

# Détecter le type de ressource selon le genre
TYPE_AUDIO_KEYWORDS = ["audio"]

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def uid():
    return str(uuid.uuid4())[:8]

def nettoyer(val):
    """Nettoie une valeur cellule."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def tronquer(texte, max_len=500):
    if not texte:
        return None
    t = str(texte).strip()
    return t[:max_len] + "…" if len(t) > max_len else t

def parser_date(val):
    if not val:
        return None
    s = str(val).strip()
    # Formats possibles : DD/MM/YYYY, YYYY, YYYY-MM-DD
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return s  # On garde tel quel si format inconnu

def parser_auteurs(val):
    if not val:
        return []
    # L'appli duplique parfois "Prénom Nom, Prénom Nom" → dédoublonner
    auteurs = [a.strip() for a in str(val).split(",") if a.strip()]
    return list(dict.fromkeys(auteurs))  # Dédoublonnage ordre préservé

def detecter_type(genres_str, type_base="livre_papier"):
    """Détecte si c'est un audio book selon le genre."""
    if not genres_str:
        return type_base
    g = genres_str.lower()
    if any(k in g for k in TYPE_AUDIO_KEYWORDS):
        return "audio"
    return type_base

def parser_genres(genres_str):
    """Extrait les genres en liste, en retirant 'Audio' (c'est un type, pas un genre)."""
    if not genres_str:
        return []
    genres = [g.strip() for g in str(genres_str).split(",") if g.strip()]
    # On garde 'Audio' comme tag mais on le signale aussi dans le type
    return genres

# ─── PARSERS PAR FEUILLE ─────────────────────────────────────────────────────

def importer_livres(ws, statut="possédé"):
    """Feuille Livres ou Livres - Liste de souhaits."""
    ressources = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        titre = nettoyer(row[0])
        if not titre:
            continue

        genres_raw = nettoyer(row[3])
        type_res = detecter_type(genres_raw, "livre_papier")

        # Colonnes liste de souhaits ont 2 colonnes de plus (Amazon, Fnac)
        if ws.max_column >= 15:  # Liste de souhaits
            lien_amazon = nettoyer(row[8])
            lien_fnac   = nettoyer(row[9])
            lu_idx, commentaire_idx, resume_idx, couv_idx = 10, 12, 13, 14
        else:
            lien_amazon = None
            lien_fnac   = None
            lu_idx, commentaire_idx, resume_idx, couv_idx = 8, 10, 11, 12

        ressource = {
            "id":           uid(),
            "type":         type_res,
            "statut":       statut,
            "titre":        titre,
            "auteurs":      parser_auteurs(row[1]),
            "serie":        nettoyer(row[2]),
            "tags":         parser_genres(genres_raw),
            "date_publication": parser_date(row[4]),
            "editeur":      nettoyer(row[5]),
            "pages":        nettoyer(row[6]),
            "isbn":         nettoyer(row[7]),
            "lu":           nettoyer(row[lu_idx]) if lu_idx < ws.max_column else None,
            "commentaire":  nettoyer(row[commentaire_idx]) if commentaire_idx < ws.max_column else None,
            "resume":       None,
            "couverture":   nettoyer(row[couv_idx]) if couv_idx < ws.max_column else None,
            "lien_amazon":  lien_amazon,
            "lien_fnac":    lien_fnac,
            # Champs spécifiques audio
            "fichier":      None,
            "narrateur":    None,
            "duree":        None,
            # Champs non remplis à ce stade
            "langue":       None,
            "localisation": None,
            "date_ajout":   datetime.today().strftime("%Y-%m-%d"),
            "source":       "import_mabibli",
        }
        ressources.append(ressource)
    return ressources


def importer_bd(ws, statut="possédé"):
    """Feuille Bandes Dessinées — même structure que Livres."""
    ressources = importer_livres(ws, statut)
    for r in ressources:
        r["type"] = "bd"
    return ressources


def importer_films(ws, statut="possédé"):
    """Feuille Films."""
    ressources = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        titre = nettoyer(row[0])
        if not titre:
            continue

        if ws.max_column >= 13:  # Liste de souhaits
            lien_amazon = nettoyer(row[8])
            lien_fnac   = nettoyer(row[9])
            vu_idx, commentaire_idx, resume_idx, couv_idx = 10, 12, 13, 14
        else:
            lien_amazon = None
            lien_fnac   = None
            vu_idx, commentaire_idx, resume_idx, couv_idx = 8, 10, 11, 12

        ressource = {
            "id":               uid(),
            "type":             "film",
            "statut":           statut,
            "titre":            titre,
            "auteurs":          parser_auteurs(row[1]),  # réalisateurs
            "serie":            nettoyer(row[2]),
            "tags":             parser_genres(row[3]),
            "date_publication": parser_date(row[4]),
            "editeur":          nettoyer(row[5]),  # société de production
            "format":           nettoyer(row[6]),
            "isbn":             nettoyer(row[7]),  # EAN
            "lu":               nettoyer(row[vu_idx]) if vu_idx < ws.max_column else None,
            "commentaire":      nettoyer(row[commentaire_idx]) if commentaire_idx < ws.max_column else None,
            "resume":       None,
            "couverture":       nettoyer(row[couv_idx]) if couv_idx < ws.max_column else None,
            "lien_amazon":      lien_amazon,
            "lien_fnac":        lien_fnac,
            "fichier":          None,
            "langue":           None,
            "localisation":     None,
            "date_ajout":       datetime.today().strftime("%Y-%m-%d"),
            "source":           "import_mabibli",
        }
        ressources.append(ressource)
    return ressources


def importer_jeux(ws, statut="possédé"):
    """Feuille Jeux Vidéo."""
    ressources = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        titre = nettoyer(row[0])
        if not titre:
            continue

        if ws.max_column >= 13:
            lien_amazon = nettoyer(row[8])
            lien_fnac   = nettoyer(row[9])
            joue_idx, commentaire_idx, resume_idx, couv_idx = 10, 12, 13, 14
        else:
            lien_amazon = None
            lien_fnac   = None
            joue_idx, commentaire_idx, resume_idx, couv_idx = 8, 10, 11, 12

        ressource = {
            "id":               uid(),
            "type":             "jeu_video",
            "statut":           statut,
            "titre":            titre,
            "auteurs":          parser_auteurs(row[1]),  # développeurs
            "serie":            nettoyer(row[2]),
            "tags":             parser_genres(row[3]),
            "date_publication": parser_date(row[4]),
            "editeur":          nettoyer(row[5]),
            "plateforme":       nettoyer(row[6]),
            "isbn":             nettoyer(row[7]),  # EAN
            "lu":               nettoyer(row[joue_idx]) if joue_idx < ws.max_column else None,
            "commentaire":      nettoyer(row[commentaire_idx]) if commentaire_idx < ws.max_column else None,
            "resume":           None,
            "couverture":       nettoyer(row[couv_idx]) if couv_idx < ws.max_column else None,
            "lien_amazon":      lien_amazon,
            "lien_fnac":        lien_fnac,
            "fichier":          None,
            "langue":           None,
            "localisation":     None,
            "date_ajout":       datetime.today().strftime("%Y-%m-%d"),
            "source":           "import_mabibli",
        }
        ressources.append(ressource)
    return ressources


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print(f"📂 Lecture de {XLSX_FILE}...")
    wb = openpyxl.load_workbook(XLSX_FILE)

    toutes_ressources = []

    # Mapping feuille → fonction d'import
    feuilles = [
        ("Livres",                          importer_livres,  "possédé"),
        ("Livres - Liste de souhaits",      importer_livres,  "souhaité"),
        ("Bandes Dessinées",                importer_bd,      "possédé"),
        ("Bandes Dessinées - Liste de sou", importer_bd,      "souhaité"),
        ("Films",                           importer_films,   "possédé"),
        ("Films - Liste de souhaits",       importer_films,   "souhaité"),
        ("Jeux Vidéo",                      importer_jeux,    "possédé"),
        ("Jeux Vidéo - Liste de souhaits",  importer_jeux,    "souhaité"),
    ]

    for nom_feuille, func, statut in feuilles:
        if nom_feuille in wb.sheetnames:
            ws = wb[nom_feuille]
            ressources = func(ws, statut)
            print(f"  ✅ {nom_feuille} : {len(ressources)} entrée(s)")
            toutes_ressources.extend(ressources)
        else:
            print(f"  ⏭  {nom_feuille} : feuille absente, ignorée")

    # Sauvegarder
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(toutes_ressources, f, ensure_ascii=False, indent=2)

    print(f"\n🎉 Import terminé : {len(toutes_ressources)} ressource(s) → {OUTPUT_FILE}")

    # Stats rapides
    from collections import Counter
    types = Counter(r["type"] for r in toutes_ressources)
    statuts = Counter(r["statut"] for r in toutes_ressources)
    print("\n📊 Répartition par type :")
    for t, n in types.most_common():
        print(f"   {t:<20} {n}")
    print("\n📊 Répartition par statut :")
    for s, n in statuts.most_common():
        print(f"   {s:<20} {n}")


if __name__ == "__main__":
    main()
